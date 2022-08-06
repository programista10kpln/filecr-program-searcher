import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

url = "https://filecr.com/category/ms-windows/"

querystring = {"page":"1"}
payload = ""
headers = {}

first_site = requests.request("GET", url, data=payload, headers=headers, params=querystring)

if first_site.status_code == 200:
    soup = BeautifulSoup(first_site.content, 'html.parser')
    max_page = int(soup.find('ul', class_='pagination').find_all('li')[-1].text) + 1
    pages = range(1, max_page)
    
    programs = []

    print('Collecting data...')

    for page in tqdm(pages):
        querystring['page'] = page
        current_site = requests.request("GET", url, data=payload, headers=headers, params=querystring)
        current_soup = BeautifulSoup(current_site.content, 'html.parser')

        products_sections = current_soup.find_all('div', class_='product')

        names = [product.find('a', class_='product-title').text.strip().replace('\n', '') if product.find('a', class_='product-title') is not None else '' for product in products_sections]
        categories = [product.find('a', class_='product-category').text.strip().replace('\n', '') if product.find('a', class_='product-category') is not None else '' for product in products_sections]
        descriptions = [product.find('p', class_='product-desc').text.strip().replace('\n', '') if product.find('p', class_='product-desc') is not None else '' for product in products_sections]
        downloads_quantities = [product.find_all('span', class_='meta-text')[-1].text.strip().replace('\n', '') if product.find_all('span', class_='meta-text') is not None else '' for product in products_sections]
        files_sizes = [product.find('div', class_='side-border product-size').text.strip().replace('\n', '') if product.find('div', class_='side-border product-size') is not None else '' for product in products_sections]
        hrefs = [product.find('a', class_='product-title')['href'] if product.find('a', class_='product-title') is not None else '' for product in products_sections]

        for name, category, description, downloads_quantity, files_size, href in zip(names, categories, descriptions, downloads_quantities, files_sizes, hrefs):
            programs.append({
                'name': name,
                'category': category, 
                'description': description, 
                'downloads_quantity': downloads_quantity, 
                'files_size': files_size,
                'href': href
            })

    # for i in programs:
    #     print(i, '\n\n')

    wb = Workbook()
    ws = wb.active

    column = 1

    for key in programs[0].keys():
        ws.cell(1, column).style = '60 % - Accent4'
        ws.cell(1, column).value = key.replace('_', ' ').capitalize()
        column += 1

    row = 2
    counter = 0

    for record in programs:
        column = 1
        for value in record.values():
            ws.cell(row, column).value = value
            column += 1
            counter += 1
            
            if counter % 6 == 0:
                row += 1
    
    # adjust columns width
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1)
        ws.column_dimensions[column].width = adjusted_width
    
    
    wb.save('programs.xlsx')
    print('Done')